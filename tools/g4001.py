"""
G4001 — 商品服務名稱初篩工具
Session State 前綴: g4001_

設計原則：
- 不使用 st.text_area（會在每次輸入時觸發 rerun）
- 不使用 st.rerun()（避免 orphan widget key 導致無限 rerun）
- 所有 widget 在每次 render 都會被渲染（不做條件隱藏 widget）
- 每個 widget 只用 key= 控制狀態，不同時傳 value=/index=
- 重量級操作（Excel 讀取）結果快取在 session state
"""

import streamlit as st
import openpyxl
import json
import io
import time
import math
from anthropic import Anthropic

# ─────────────────────────── 常數 ───────────────────────────
_PREFIX = "g4001_"
_BATCH_SIZE = 150
_MODEL = "claude-sonnet-4-20250514"


# ─────────────────────────── 內部函式 ───────────────────────────

def _init_session_state():
    defaults = {
        f"{_PREFIX}uploaded_name": None,
        f"{_PREFIX}uploaded_bytes": None,
        f"{_PREFIX}file_fingerprint": None,
        f"{_PREFIX}sheet_names": [],
        f"{_PREFIX}sheet_max_rows": {},
        f"{_PREFIX}is_processing": False,
        f"{_PREFIX}results": None,
        f"{_PREFIX}output_bytes": None,
        f"{_PREFIX}output_name": None,
        f"{_PREFIX}elapsed_seconds": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _detect_category_sheets(wb):
    sheets = []
    max_rows = {}
    for name in wb.sheetnames:
        s = name.strip()
        if s.isdigit():
            sheets.append(s)
            max_rows[s] = wb[s].max_row or 2
    return sheets, max_rows


def _read_sheet_items(ws):
    items = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name:
            continue
        name_en = ws.cell(r, 4).value or ""
        items.append({
            "row": r,
            "name": str(name).strip(),
            "name_en": str(name_en).strip(),
        })
    return items


def _build_prompt(items_batch, basis_text, strictness, sheet_name, mode):
    strictness_map = {
        "寧寬勿嚴（多標）":
            "判斷標準請寧寬勿嚴。只要有合理的商業關聯可能性就標記。灰色地帶的項目一律標記。",
        "適中":
            "判斷標準適中。明確相關的標記，明確無關的不標記，灰色地帶的項目依據一般商業常識合理判斷。",
        "寧嚴勿寬（少標）":
            "判斷標準請寧嚴勿寬。只標記非常明確屬於該類型品牌會直接經營、販售的商品或服務。灰色地帶的項目一律不標記。",
    }
    items_text = "\n".join(
        f"- {it['name']}" + (f" ({it['name_en']})" if it['name_en'] else "")
        for it in items_batch
    )

    if mode == "品牌模式":
        task_text = "你是商標分類專家。請根據以下比對基礎，判斷清單中哪些商品或服務項目屬於該類型品牌會經營、販售的商品或服務。"
    else:
        task_text = "你是商標分類專家。請根據以下比對基礎中列出的商品或服務名稱，從清單中找出屬於同義詞、近似詞、或在商業上常見於同一店面或賣場販售的商品或服務項目。"

    return f"""{task_text}

## 比對基礎
{basis_text}

## 判斷寬嚴度
{strictness_map.get(strictness, strictness_map["適中"])}

## 目前處理的類別
第 {sheet_name} 類

## 待判斷的商品/服務項目清單
{items_text}

## 回傳格式要求
請只回傳一個 JSON 物件，不要加任何 markdown 標記或其他文字。
格式為: {{"marked": ["項目名稱1", "項目名稱2", ...]}}
其中 marked 陣列包含你認為應該標記的項目名稱（必須與上方清單中的名稱完全一致）。
如果這批項目中沒有任何項目應該被標記，回傳: {{"marked": []}}"""


def _build_prompt_preview(basis_text, strictness, mode):
    """產生 prompt 前段內容的預覽（不含項目清單和回傳格式要求）。"""
    strictness_map = {
        "寧寬勿嚴（多標）":
            "判斷標準請寧寬勿嚴。只要有合理的商業關聯可能性就標記。灰色地帶的項目一律標記。",
        "適中":
            "判斷標準適中。明確相關的標記，明確無關的不標記，灰色地帶的項目依據一般商業常識合理判斷。",
        "寧嚴勿寬（少標）":
            "判斷標準請寧嚴勿寬。只標記非常明確屬於該類型品牌會直接經營、販售的商品或服務。灰色地帶的項目一律不標記。",
    }
    if mode == "品牌模式":
        task_text = "你是商標分類專家。請根據以下比對基礎，判斷清單中哪些商品或服務項目屬於該類型品牌會經營、販售的商品或服務。"
    else:
        task_text = "你是商標分類專家。請根據以下比對基礎中列出的商品或服務名稱，從清單中找出屬於同義詞、近似詞、或在商業上常見於同一店面或賣場販售的商品或服務項目。"

    return f"""{task_text}

## 比對基礎
{basis_text}

## 判斷寬嚴度
{strictness_map.get(strictness, strictness_map["適中"])}"""


def _build_refine_prompt(marked_items, basis_text, max_count, sheet_name):
    items_text = "\n".join(
        f"- {it['name']}" + (f" ({it['name_en']})" if it['name_en'] else "")
        for it in marked_items
    )
    return f"""你是商標分類專家。以下是第 {sheet_name} 類中，已初步判斷為符合條件的商品/服務項目清單（共 {len(marked_items)} 項）。

## 比對基礎
{basis_text}

## 任務
這些項目都與上述比對基礎有關聯，但客戶希望每個類別最多只選擇 {max_count} 個最相關的項目。
請從以下清單中，挑選出與比對基礎最直接相關、最核心的 {max_count} 個項目。

## 已標記的項目清單
{items_text}

## 回傳格式要求
請只回傳一個 JSON 物件，不要加任何 markdown 標記或其他文字。
格式為: {{"marked": ["項目名稱1", "項目名稱2", ...]}}
marked 陣列中最多 {max_count} 個項目，名稱必須與上方清單完全一致。"""


def _call_claude_api(api_key, prompt):
    client = Anthropic(api_key=api_key)
    try:
        resp = client.messages.create(
            model=_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[1] if "\n" in text else text[3:]
            if text.endswith("```"):
                text = text[:-3]
            text = text.strip()
        return json.loads(text)
    except json.JSONDecodeError:
        return None
    except Exception as e:
        st.error(f"Claude API 呼叫失敗: {e}")
        return None


def _process_sheet(ws, sheet_name, api_key, basis_text, strictness, mode,
                   progress_bar, status_text, prog_base, prog_weight):
    items = _read_sheet_items(ws)
    if not items:
        return 0, 0, set(), []

    total = len(items)
    n_batches = math.ceil(total / _BATCH_SIZE)
    marked_names = set()

    for bi in range(n_batches):
        s = bi * _BATCH_SIZE
        e = min(s + _BATCH_SIZE, total)
        batch = items[s:e]
        status_text.text(f"第 {sheet_name} 類：正在判斷第 {s+1}～{e} 項（共 {total} 項）…")
        prompt = _build_prompt(batch, basis_text, strictness, sheet_name, mode)
        result = _call_claude_api(api_key, prompt)
        if result and "marked" in result:
            for name in result["marked"]:
                marked_names.add(name.strip())
        progress_bar.progress(min(prog_base + prog_weight * ((bi + 1) / n_batches), 1.0))
        if bi < n_batches - 1:
            time.sleep(1)

    marked_count = 0
    for item in items:
        if item["name"] in marked_names:
            ws.cell(item["row"], 5).value = "O"
            marked_count += 1
        else:
            ws.cell(item["row"], 5).value = None

    return marked_count, total, marked_names, items


def _refine_sheet(ws, items, marked_names, api_key, basis_text,
                  max_count, sheet_name, status_text):
    marked_items = [it for it in items if it["name"] in marked_names]
    status_text.text(
        f"第 {sheet_name} 類：已標記 {len(marked_items)} 項，超過上限 {max_count} 項，"
        f"正在由 AI 篩選最相關的 {max_count} 項…"
    )
    prompt = _build_refine_prompt(marked_items, basis_text, max_count, sheet_name)
    result = _call_claude_api(api_key, prompt)
    refined = set()
    if result and "marked" in result:
        for name in result["marked"]:
            refined.add(name.strip())
    if not refined:
        return len(marked_items)
    count = 0
    for item in items:
        if item["name"] in refined:
            ws.cell(item["row"], 5).value = "O"
            count += 1
        else:
            ws.cell(item["row"], 5).value = None
    return count


# ─────────────────────────── render 入口 ───────────────────────────

def render(api_key=None):
    _init_session_state()

    # ──── DEBUG: 在側邊欄顯示每個步驟的執行狀態 ────
    _dbg = st.sidebar.expander("🐛 DEBUG LOG", expanded=True)
    _t0 = time.time()
    def _dbg_log(msg):
        elapsed_ms = int((time.time() - _t0) * 1000)
        _dbg.write(f"`{elapsed_ms:>5}ms` {msg}")
    _dbg_log("render() 開始")

    st.markdown(
        '<div class="step-header">'
        '<span class="step-badge">G4001</span>'
        '商品服務名稱初篩工具'
        '</div>',
        unsafe_allow_html=True,
    )
    st.caption("上傳商品服務名稱清單 Excel，由 AI 根據您指定的比對基礎自動標示相關項目。")

    # ════════════════════════════════════════
    # 步驟 1：上傳 Excel
    # ════════════════════════════════════════
    st.markdown("#### 步驟 1：上傳商品服務名稱清單表")
    uploaded = st.file_uploader("選擇 .xlsx 檔案", type=["xlsx"],
                                key=f"{_PREFIX}uploader")

    _dbg_log(f"file_uploader 完成, uploaded={'有檔案' if uploaded else 'None'}")

    if uploaded is not None:
        fp = uploaded.file_id
        old_fp = st.session_state[f"{_PREFIX}file_fingerprint"]
        _dbg_log(f"file_id={fp}, 舊fingerprint={old_fp}, 相同={fp == old_fp}")
        if fp != old_fp:
            _dbg_log("fingerprint 不同，開始讀取 Excel…")
            raw = uploaded.getvalue()
            _dbg_log(f"getvalue() 完成, {len(raw)} bytes")
            st.session_state[f"{_PREFIX}uploaded_bytes"] = raw
            st.session_state[f"{_PREFIX}uploaded_name"] = uploaded.name
            st.session_state[f"{_PREFIX}file_fingerprint"] = fp
            st.session_state[f"{_PREFIX}results"] = None
            st.session_state[f"{_PREFIX}output_bytes"] = None
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
                _dbg_log("load_workbook 完成")
                sheets, max_rows = _detect_category_sheets(wb)
                _dbg_log(f"偵測到類別頁: {sheets}, max_rows: {max_rows}")
                wb.close()
                st.session_state[f"{_PREFIX}sheet_names"] = sheets
                st.session_state[f"{_PREFIX}sheet_max_rows"] = max_rows
            except Exception as e:
                _dbg_log(f"ERROR: {e}")
                st.error(f"無法讀取 Excel 檔案: {e}")
                return
            _dbg_log("Excel 處理完成")
        else:
            _dbg_log("fingerprint 相同，跳過讀取")
    else:
        _dbg_log("沒有上傳檔案")

    if not st.session_state[f"{_PREFIX}uploaded_bytes"]:
        st.info("請上傳商品服務名稱清單表（.xlsx 格式）。")
        _dbg_log("uploaded_bytes 為空，return")
        return

    detected = st.session_state[f"{_PREFIX}sheet_names"]
    if not detected:
        st.warning("未偵測到任何以數字命名的類別頁。")
        _dbg_log("無類別頁，return")
        return

    st.success(f"已偵測到 {len(detected)} 個類別頁：{', '.join(detected)}")
    _dbg_log(f"顯示 {len(detected)} 個類別頁")

    # ════════════════════════════════════════
    # 步驟 2：選擇要檢查的類別
    # ════════════════════════════════════════
    st.markdown("#### 步驟 2：選擇要檢查的類別")
    st.caption("勾選要處理的類別（預設全選）")
    selected = []
    cols = st.columns(min(len(detected), 6))
    for idx, sn in enumerate(detected):
        ck = f"{_PREFIX}cb_{sn}"
        if ck not in st.session_state:
            st.session_state[ck] = True
        with cols[idx % len(cols)]:
            if st.checkbox(f"第 {sn} 類", key=ck):
                selected.append(sn)

    _dbg_log(f"步驟2 完成, selected={selected}")

    if not selected:
        st.warning("請至少選擇一個類別。")
        _dbg_log("無選擇，return")
        return

    # ════════════════════════════════════════
    # 步驟 3：選擇比對模式 + 輸入比對基礎
    # ════════════════════════════════════════
    st.markdown("#### 步驟 3：輸入比對基礎")
    if f"{_PREFIX}mode_r" not in st.session_state:
        st.session_state[f"{_PREFIX}mode_r"] = "品牌模式"
    mode = st.radio(
        "比對模式",
        ["品牌模式", "商品或服務模式"],
        horizontal=True,
        key=f"{_PREFIX}mode_r",
        help="品牌模式：輸入品牌類型描述，判斷該品牌會經營什麼。\n\n商品或服務模式：輸入具體商品/服務名稱，找出同義詞、近似詞及商業上相關的項目。",
    )

    if mode == "品牌模式":
        st.markdown(
            "描述品牌類型或是申請人的經營項目，越具體越好  \n"
            "範例：連鎖手搖飲品牌，參考 CoCo都可、50嵐、迷客夏、一沐日、春水堂、得正、五桐號"
        )
        placeholder = "例：連鎖手搖飲品牌，參考 CoCo都可、50嵐"
    else:
        st.markdown(
            "輸入具體的商品名稱和／或服務名稱，用頓號或逗號分隔  \n"
            "範例：鋼筆、蠟筆、原子筆、彩色筆、水彩、顏料、西卡紙、宣紙"
        )
        placeholder = "例：鋼筆、蠟筆、原子筆、彩色筆、水彩、顏料"

    if f"{_PREFIX}basis_input" not in st.session_state:
        st.session_state[f"{_PREFIX}basis_input"] = ""
    basis = st.text_input("比對基礎", key=f"{_PREFIX}basis_input",
                          label_visibility="collapsed",
                          placeholder=placeholder)

    _dbg_log(f"步驟3 text_input 完成, basis='{basis[:30]}...' (len={len(basis)})")

    if not basis.strip():
        st.info("請輸入比對基礎描述後按 Enter。")
        _dbg_log("basis 為空，return")
        return

    # ════════════════════════════════════════
    # 步驟 4：判斷寬嚴度
    # ════════════════════════════════════════
    st.markdown("#### 步驟 4：判斷寬嚴度")
    opts = ["寧寬勿嚴（多標）", "適中", "寧嚴勿寬（少標）"]
    if f"{_PREFIX}strictness_r" not in st.session_state:
        st.session_state[f"{_PREFIX}strictness_r"] = "適中"
    strictness = st.radio("判斷傾向", opts, horizontal=True,
                          key=f"{_PREFIX}strictness_r",
                          label_visibility="collapsed")

    _dbg_log(f"步驟4 完成, strictness={strictness}")

    # ════════════════════════════════════════
    # 步驟 5：每類標記上限（選填）
    # ════════════════════════════════════════
    st.markdown("#### 步驟 5：每類標記上限（選填）")
    if f"{_PREFIX}limit_cb" not in st.session_state:
        st.session_state[f"{_PREFIX}limit_cb"] = False
    limit_on = st.checkbox("啟用每類標記上限（超過時由 AI 挑選最相關的項目）",
                           key=f"{_PREFIX}limit_cb")
    max_per_class = None
    if limit_on:
        if f"{_PREFIX}limit_sl" not in st.session_state:
            st.session_state[f"{_PREFIX}limit_sl"] = 30
        max_per_class = st.slider("每類最多標記幾項", 1, 100, step=1,
                                  key=f"{_PREFIX}limit_sl")

    _dbg_log(f"步驟5 完成, limit_on={limit_on}, max_per_class={max_per_class}")

    # ════════════════════════════════════════
    # API Key 檢查
    # ════════════════════════════════════════
    if not api_key:
        st.warning("請在側邊欄輸入 Claude API Key 後才能執行比對。")
        _dbg_log("無 API Key，return")
        return

    _dbg_log("API Key 存在")

    # ════════════════════════════════════════
    # Prompt 預覽
    # ════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 📋 Prompt 預覽（傳送給 Claude AI 的指示前段）")
    preview = _build_prompt_preview(basis, strictness, mode)
    st.code(preview, language=None)

    # ════════════════════════════════════════
    # 開始比對按鈕 + 預估費用
    # ════════════════════════════════════════
    st.markdown("---")

    if st.session_state[f"{_PREFIX}is_processing"]:
        st.warning("正在處理中，請等待完成…")
        _dbg_log("is_processing=True，return")
        return

    max_rows = st.session_state[f"{_PREFIX}sheet_max_rows"]
    total_items = sum(max(max_rows.get(sn, 2) - 2, 0) for sn in selected)
    est_calls = math.ceil(total_items / _BATCH_SIZE) if total_items > 0 else 0
    est_usd = est_calls * (1500 * 3 / 1_000_000 + 500 * 15 / 1_000_000)
    est_ntd = est_usd * 33

    col_btn, col_info = st.columns([1, 3])
    with col_btn:
        go = st.button("🚀 開始比對", key=f"{_PREFIX}go_btn", type="primary")
    with col_info:
        st.caption(
            f"約 {total_items} 個項目，預估呼叫 Claude API **{est_calls}** 次\n\n"
            f"預估費用：約 NT${est_ntd:.1f}（US${est_usd:.3f}）"
        )

    _dbg_log(f"按鈕區 完成, go={go}")
    _dbg_log("render() 結束（未按開始比對）")

    # ════════════════════════════════════════
    # 執行比對
    # ════════════════════════════════════════
    if go:
        st.session_state[f"{_PREFIX}is_processing"] = True
        st.session_state[f"{_PREFIX}results"] = None
        st.session_state[f"{_PREFIX}output_bytes"] = None

        t0 = time.time()
        wb = openpyxl.load_workbook(
            io.BytesIO(st.session_state[f"{_PREFIX}uploaded_bytes"]))

        bar = st.progress(0.0)
        status = st.empty()
        summary = {}
        n = len(selected)

        for i, sn in enumerate(selected):
            if sn not in wb.sheetnames:
                st.warning(f"工作頁 '{sn}' 不存在，跳過。")
                continue
            ws = wb[sn]
            mc, tc, mn, items = _process_sheet(
                ws, sn, api_key, basis, strictness, mode,
                bar, status, i / n, 1.0 / n)
            if max_per_class is not None and mc > max_per_class:
                mc = _refine_sheet(ws, items, mn, api_key, basis,
                                   max_per_class, sn, status)
            summary[sn] = {"marked": mc, "total": tc}

        bar.progress(1.0)
        status.text("全部完成！正在產生輸出檔案…")

        out = io.BytesIO()
        wb.save(out)
        wb.close()

        orig = st.session_state[f"{_PREFIX}uploaded_name"] or "output.xlsx"
        oname = (orig[:-5] + "_標示結果.xlsx") if orig.endswith(".xlsx") else (orig + "_標示結果.xlsx")

        st.session_state[f"{_PREFIX}results"] = summary
        st.session_state[f"{_PREFIX}output_bytes"] = out.getvalue()
        st.session_state[f"{_PREFIX}output_name"] = oname
        st.session_state[f"{_PREFIX}elapsed_seconds"] = time.time() - t0
        st.session_state[f"{_PREFIX}is_processing"] = False
        status.text("✅ 比對完成！")

    # ════════════════════════════════════════
    # 顯示結果
    # ════════════════════════════════════════
    if st.session_state[f"{_PREFIX}results"]:
        st.markdown("---")
        elapsed = st.session_state.get(f"{_PREFIX}elapsed_seconds")
        if elapsed is not None:
            m, s = int(elapsed // 60), int(elapsed % 60)
            st.markdown(f"#### 比對結果摘要　　（本次初篩耗時：{m} 分 {s:02d} 秒）")
        else:
            st.markdown("#### 比對結果摘要")

        res = st.session_state[f"{_PREFIX}results"]
        for sn, info in res.items():
            mk, tt = info["marked"], info["total"]
            pct = f"{mk/tt*100:.1f}%" if tt > 0 else "0%"
            st.markdown(
                f'<div class="result-item">'
                f'第 <strong>{sn}</strong> 類：共 {tt} 項，標記 <strong>{mk}</strong> 項（{pct}）'
                f'</div>', unsafe_allow_html=True)

        st.markdown(
            f"**合計：{sum(v['total'] for v in res.values())} 項中"
            f"標記了 {sum(v['marked'] for v in res.values())} 項**")

    # ════════════════════════════════════════
    # 下載按鈕
    # ════════════════════════════════════════
    if st.session_state[f"{_PREFIX}output_bytes"]:
        st.markdown("---")
        st.download_button(
            "📥 下載標示結果 Excel",
            data=st.session_state[f"{_PREFIX}output_bytes"],
            file_name=st.session_state[f"{_PREFIX}output_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{_PREFIX}dl_btn")
        st.caption("標示結果已寫入各類別頁的欄位 E（符合者標示 \"O\"，不符合者清空）。")
